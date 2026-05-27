"""从聊天附件中提取可读文本（仅用于会话上下文，不入库）。"""
import base64
import os
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from io import BytesIO
from typing import Optional

_W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def extract_text_from_upload(filename: str, content_b64: str, max_chars: int = 120_000) -> Optional[str]:
    """将 base64 编码的上传文件解析为纯文本。"""
    if not filename or not content_b64:
        return None

    try:
        data = base64.b64decode(content_b64)
    except Exception:
        return None

    if not data:
        return None

    name_lower = filename.lower()

    if name_lower.endswith((".txt", ".md", ".csv", ".json", ".xml", ".html", ".htm", ".log")):
        text = data.decode("utf-8", errors="ignore").strip()
        return _truncate(text, max_chars) or None

    if name_lower.endswith(".docx"):
        text = _extract_docx_text(data).strip()
        return _truncate(text, max_chars) or None

    if name_lower.endswith((".pdf", ".doc", ".xlsx", ".xls")):
        suffix = os.path.splitext(filename)[1] or ".bin"
        tmp_path = ""
        try:
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                tmp.write(data)
                tmp_path = tmp.name
            text = _extract_document_text(tmp_path, name_lower).strip()
            return _truncate(text, max_chars) or None
        except Exception:
            return None
        finally:
            if tmp_path and os.path.exists(tmp_path):
                os.unlink(tmp_path)

    text = data.decode("utf-8", errors="ignore").strip()
    return _truncate(text, max_chars) or None


def _truncate(text: str, max_chars: int) -> str:
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + "\n\n...(内容过长，已截断)..."


def _extract_docx_text(data: bytes) -> str:
    with zipfile.ZipFile(BytesIO(data)) as zf:
        with zf.open("word/document.xml") as xml_file:
            root = ET.parse(xml_file).getroot()

    paragraphs: list[str] = []
    for paragraph in root.iter(f"{_W_NS}p"):
        parts: list[str] = []
        for node in paragraph.iter():
            if node.tag == f"{_W_NS}t" and node.text:
                parts.append(node.text)
            elif node.tag == f"{_W_NS}tab":
                parts.append("\t")
        text = "".join(parts).strip()
        if text:
            paragraphs.append(text)
    return "\n\n".join(paragraphs)


def _extract_document_text(file_path: str, name_lower: str) -> str:
    if name_lower.endswith(".pdf"):
        return _extract_pdf_text(file_path)
    if name_lower.endswith(".doc"):
        return _extract_legacy_word_text(file_path)
    if name_lower.endswith((".xlsx", ".xls")):
        return _extract_excel_text(file_path)
    raise ValueError(f"不支持的文件类型: {file_path}")


def _extract_legacy_word_text(file_path: str) -> str:
    from langchain_community.document_loaders import Docx2txtLoader

    raw_docs = Docx2txtLoader(file_path).load()
    parts = [(doc.page_content or "").strip() for doc in raw_docs]
    return "\n\n".join(part for part in parts if part)


def _extract_pdf_text(file_path: str) -> str:
    try:
        from pypdf import PdfReader

        reader = PdfReader(file_path)
        parts = [(page.extract_text() or "").strip() for page in reader.pages]
        return "\n\n".join(part for part in parts if part)
    except ImportError:
        from langchain_community.document_loaders import PyPDFLoader

        raw_docs = PyPDFLoader(file_path).load()
        parts = [(doc.page_content or "").strip() for doc in raw_docs]
        return "\n\n".join(part for part in parts if part)


def _extract_excel_text(file_path: str) -> str:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheets: list[str] = []
        for sheet in workbook.worksheets:
            rows: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if cells:
                    rows.append("\t".join(cells))
            if rows:
                sheets.append(f"[{sheet.title}]\n" + "\n".join(rows))
        return "\n\n".join(sheets)
    except ImportError:
        from langchain_community.document_loaders import UnstructuredExcelLoader

        raw_docs = UnstructuredExcelLoader(file_path).load()
        parts = [(doc.page_content or "").strip() for doc in raw_docs]
        return "\n\n".join(part for part in parts if part)
